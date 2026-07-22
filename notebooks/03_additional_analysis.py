import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from scipy.optimize import minimize
from scipy.stats import norm, linregress
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows
from scipy.optimize import OptimizeWarning
import warnings

# Suprimir warnings del optimizador SLSQP
warnings.filterwarnings("ignore", category=RuntimeWarning)
warnings.filterwarnings("ignore", category=OptimizeWarning)

# === 1. Leer y combinar hojas del Excel ===
input_path = "rentabilidades_acciones.xlsx"
xls = pd.ExcelFile(input_path)
returns = pd.DataFrame()
for sheet in xls.sheet_names:
    df = xls.parse(sheet)
    # Convertir Date a índice datetime
    df["Date"] = pd.to_datetime(df["Date"])
    df = df.set_index("Date")
    # Convertir strings “1,23%” a floats 0.0123
    for col in df.columns:
        if df[col].dtype == object or df[col].dtype == str:
            df[col] = (
                df[col]
                .astype(str)
                .str.replace("%","")
                .str.replace(",","." )
                .astype(float)
                / 100
            )
    # Renombrar columnas con prefijo de país
    df.columns = [f"{sheet}_{col}" for col in df.columns]
    returns = pd.concat([returns, df], axis=1)

# === 2. Cálculo de Betas por país ===
beta_rows = []
for sheet in xls.sheet_names:
    # Extraer las columnas de ese país
    cols = [c for c in returns.columns if c.startswith(f"{sheet}_")]
    df = returns[cols].dropna(how="all")
    if df.empty:
        continue
    # Mercado proxy = retorno promedio de todos los activos de ese país
    mkt = df.mean(axis=1)
    for asset in cols:
        # emparejar índices no-NaN
        common = mkt.dropna().index.intersection(df[asset].dropna().index)
        if len(common) < 2:
            continue
        slope, _, _, _, _ = linregress(
            mkt.loc[common],
            df[asset].loc[common]
        )
        beta_rows.append({
            "Pais": sheet,
            "Activo": asset,
            "Beta": slope
        })

betas_df = pd.DataFrame(beta_rows)
# Guardar Betas
betas_df.to_excel("betas_por_pais.xlsx", index=False)
print("✅ Betas guardadas en betas_por_pais.xlsx")

# === 3. Parámetro libre de riesgo ===
rf = 0.0034

# === 4. Funciones de optimización ===
def performance(w, mu, Sigma):
    ret = w @ mu
    risk = np.sqrt(w @ Sigma @ w)
    return ret, risk

def min_var(mu, Sigma):
    n = len(mu)
    obj = lambda w: np.sqrt(w @ Sigma @ w)
    cons = {'type':'eq','fun':lambda w: np.sum(w)-1}
    bounds = [(0.01,0.10)]*n
    w0 = np.ones(n)/n
    return minimize(obj, w0, method="SLSQP", bounds=bounds, constraints=cons).x

def max_ret(mu, Sigma):
    n = len(mu)
    obj = lambda w: -(w @ mu)
    cons = {'type':'eq','fun':lambda w: np.sum(w)-1}
    bounds = [(0.01,0.10)]*n
    w0 = np.ones(n)/n
    return minimize(obj, w0, method="SLSQP", bounds=bounds, constraints=cons).x

def max_sharpe(mu, Sigma):
    n = len(mu)
    obj = lambda w: -((w @ mu - rf)/np.sqrt(w @ Sigma @ w))
    cons = {'type':'eq','fun':lambda w: np.sum(w)-1}
    bounds = [(0.01,0.10)]*n
    w0 = np.ones(n)/n
    return minimize(obj, w0, method="SLSQP", bounds=bounds, constraints=cons).x

# === 5. Generación Frontera Eficiente ===
def generar_grafico(mu, Sigma, w_min, w_max, w_sh, imgfile):
    # Puntos extremos
    ret_min, risk_min = performance(w_min, mu, Sigma)
    ret_max, risk_max = performance(w_max, mu, Sigma)
    ret_sh, risk_sh = performance(w_sh, mu, Sigma)
    sr = (ret_sh - rf)/risk_sh
    # 30 intermedios
    R_t = np.linspace(ret_min, ret_max, 30)
    risks = []; weights=[]
    n = len(mu)
    for r in R_t:
        cons = [
            {'type':'eq','fun':lambda w: np.sum(w)-1},
            {'type':'eq','fun':lambda w: w@mu - r}
        ]
        bnds = [(0.01,0.10)]*n
        w0 = np.ones(n)/n
        w = minimize(lambda w: np.sqrt(w@Sigma@w), w0,
                     method="SLSQP", bounds=bnds, constraints=cons).x
        weights.append(w)
        risks.append(np.sqrt(w@Sigma@w))
    df_front = pd.DataFrame(weights, columns=mu.index).T
    df_front.columns = [f"PF_{i+1}" for i in range(30)]
    df_front.loc["Rentabilidad"] = R_t
    df_front.loc["Riesgo"] = risks

    # Plot
    plt.figure(figsize=(9,5))
    plt.plot(risks, R_t, 'b-', label="Frontera")
    plt.scatter(risks, R_t, s=15)
    plt.scatter(risk_min, ret_min, c='g', label="Min Var")
    plt.scatter(risk_max, ret_max, c='r', label="Max Ret")
    plt.scatter(risk_sh, ret_sh, c='orange', marker='*', s=200, label="Max Sharpe")
    x = np.linspace(0, max(risks)*1.1, 100)
    plt.plot(x, rf + sr*x, 'k--', label="CML")
    plt.xlabel("Riesgo"); plt.ylabel("Retorno")
    plt.legend(); plt.grid(True); plt.tight_layout()
    plt.savefig(imgfile, dpi=300); plt.close()

    return df_front

# === 6. Guardar en Excel ===
def guardar_excel(fname, df_w, df_sum, Sigma, df_front=None, img=None):
    with pd.ExcelWriter(fname, engine="openpyxl") as writer:
        df_w.to_excel(writer, "Pesos")
        df_sum.to_excel(writer, "Resumen", index=False)
        Sigma.to_excel(writer, "Covarianzas")
        if df_front is not None:
            df_front.to_excel(writer, "Pesos Frontera")
    if img:
        wb = load_workbook(fname)
        ws = wb.create_sheet("Grafico", 0)
        pic = XLImage(img); pic.anchor="A1"; ws.add_image(pic)
        wb.save(fname)

# === 7. Análisis Portafolios ===
def analizar(df, xlsx, png, plot_front=False, save_front=False):
    mu = df.mean(); Sigma=df.cov()
    w_min, w_max, w_sh = min_var(mu,Sigma), max_ret(mu,Sigma), max_sharpe(mu,Sigma)
    df_w = pd.DataFrame({
        "MinVar":w_min,"MaxRet":w_max,"MaxSharpe":w_sh
    }, index=df.columns)
    df_w_fmt = df_w.applymap(lambda x:f"{x*100:.2f}%")
    # resumen
    ret_min,r_min=performance(w_min,mu,Sigma)
    ret_max,r_max=performance(w_max,mu,Sigma)
    ret_sh,r_sh=performance(w_sh,mu,Sigma)
    df_sum=pd.DataFrame({
        "Port": ["MinVar","MaxRet","MaxSharpe"],
        "Ret(%)":[ret_min*100, ret_max*100, ret_sh*100],
        "Risk(%)":[r_min*100, r_max*100, r_sh*100],
        "Sharpe":["","", (ret_sh-rf)/r_sh]
    })
    df_front=None
    if plot_front:
        df_front=generar_grafico(mu,Sigma,w_min,w_max,w_sh,png)
    else:
        # solo puntos + CML
        ret_sh,r_sh=performance(w_sh,mu,Sigma)
        sr=(ret_sh-rf)/r_sh
        plt.figure(figsize=(6,4))
        plt.scatter(r_min,ret_min,c="g",label="MinVar")
        plt.scatter(r_max,ret_max,c="r",label="MaxRet")
        plt.scatter(r_sh,ret_sh,c="orange",marker="*",s=150,label="MaxSharpe")
        x=np.linspace(0,max(r_min,r_max,r_sh)*1.1,50)
        plt.plot(x,rf+sr*x,"k--",label="CML")
        plt.legend(); plt.grid(True); plt.tight_layout()
        plt.savefig(png,dpi=300); plt.close()
    guardar_excel(xlsx, df_w_fmt, df_sum, Sigma, df_front if save_front else None, png)
    print(f"Guardado: {xlsx}")

# ejecutar total, conservador y agresivo
analizar(returns, "resultados_portafolios.xlsx","fr_total.png",
         plot_front=True, save_front=True)
risks = np.sqrt(np.diag(returns.cov()))
serie = pd.Series(risks,index=returns.columns).sort_values()
analizar(returns[serie.head(35).index],
         "portafolio_conservador.xlsx","fr_cons.png",
         plot_front=True, save_front=False)
analizar(returns[serie.tail(35).index],
         "portafolio_agresivo.xlsx","fr_agr.png",
         plot_front=True, save_front=False)

# === 8. Inciso III: VaR 70% + PDF/CDF ===
mu=returns.mean(); Sigma=returns.cov()
w_sh=max_sharpe(mu,Sigma)
mu_p,sig_p=performance(w_sh,mu,Sigma)
z70=norm.ppf(0.70)
VaR_param=mu_p - z70*sig_p
rt=returns@w_sh; sorted_rt=np.sort(rt)
VaR_hist=-sorted_rt[int(np.ceil(0.30*len(sorted_rt)))-1]
sim=np.random.normal(mu_p,sig_p,10000)
VaR_mc=-np.sort(sim)[int(np.ceil(0.30*10000))-1]
df_var=pd.DataFrame({
    "Metodo":["Parametrico","Historico","MonteCarlo"],
    "VaR70(%)":[VaR_param*100,VaR_hist*100,VaR_mc*100]
})
df_par=pd.DataFrame({"Par":["mu_p","sigma_p"],"Val":[mu_p,sig_p]})
with pd.ExcelWriter("inciso3.xlsx",engine="openpyxl") as w:
    df_var.to_excel(w,"VaR_70",index=False)
    df_par.to_excel(w,"Portafolio",index=False)
# PDF/CDF
x=np.linspace(mu_p-4*sig_p,mu_p+4*sig_p,200)
plt.figure(); plt.plot(x,norm.pdf(x,mu_p,sig_p)); plt.title("PDF"); plt.grid(); plt.tight_layout(); plt.savefig("inc3_pdf.png"); plt.close()
plt.figure(); plt.plot(x,norm.cdf(x,mu_p,sig_p)); plt.title("CDF"); plt.grid(); plt.tight_layout(); plt.savefig("inc3_cdf.png"); plt.close()

# === 9. Inciso V: Portafolio Beta objetivo ===
# recorte 2017-01 a 2024-12
trim=returns.loc["2017-01-01":"2024-12-31"]
# definir proxies (ajustar nombres según tus hojas)
proxies = {sheet: trim[f"{sheet}_MktIdx"] for sheet in xls.sheet_names if f"{sheet}_MktIdx" in trim}
# activos = todo menos proxies
assets = [c for c in trim.columns if c not in proxies.values()]
# betas (ya calculadas antes, pero recalculamos por si acaso)
betas = {}
for a in assets:
    country=a.split("_")[0]
    m = proxies[country].dropna()
    r = trim[a].loc[m.index]
    if len(r)>1:
        betas[a] = linregress(m, r).slope
betas = pd.Series(betas)

beta_tgt=1.20
n=len(assets)
# función objetivo y restricciones
obj_b = lambda w: np.sqrt(w @ Sigma.loc[assets,assets] @ w)
cons_b = [
    {'type':'eq','fun':lambda w: np.sum(w)-1},
    {'type':'eq','fun':lambda w: w@betas.values-beta_tgt}
]
bnds=[(0.01,0.10)]*n
w0=np.ones(n)/n
res= minimize(obj_b,w0,method="SLSQP",bounds=bnds,constraints=cons_b)
w_b=res.x
mu_b,sig_b = performance(w_b, mu[assets], Sigma.loc[assets,assets])

# guardar inciso5.xlsx
wb=Workbook(); ws1=wb.active; ws1.title="PesosBeta"
dfb=pd.DataFrame({'Activo':assets,'Peso':w_b,'Beta':betas.values})
for r in dataframe_to_rows(dfb,index=False,header=True): ws1.append(r)
ws2=wb.create_sheet("ResumenBeta")
dfb2=pd.DataFrame({
    "Parametro":["Beta_obj","Beta_eff","Retorno","Riesgo"],
    "Valor":[beta_tgt, w_b@betas, mu_b, sig_b]
})
for r in dataframe_to_rows(dfb2,index=False,header=True): ws2.append(r)
wb.save("inciso5.xlsx")
print("✅ Proceso completo")
